import os
import ast

import openpyxl
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import pyrebase


def upload_to_drive(filename):
    gauth = GoogleAuth()
    # Try to load saved client credentials
    gauth.LoadCredentialsFile("oauthcreds.txt")

    if gauth.credentials is None:
        # Authenticate if they're not there

        # https://stackoverflow.com/a/55876179
        # This is what solved the issues:
        gauth.GetFlow()
        gauth.flow.params.update({'access_type': 'offline'})
        gauth.flow.params.update({'approval_prompt': 'force'})

        gauth.LocalWebserverAuth()

    elif gauth.access_token_expired:
        # Refresh them if expired
        gauth.Refresh()

    else:
        # Initialize the saved creds
        gauth.Authorize()
    # Save the current credentials to a file
    gauth.SaveCredentialsFile("oauthcreds.txt")

    drive = GoogleDrive(gauth)

    # https://stackoverflow.com/a/22934892
    # https://stackoverflow.com/a/40236586
    drive_file = drive.CreateFile({'title': filename,
                                   "parents": [{"kind": "drive#fileLink", "id": os.environ['drive_file_id']}]})

    # Read file and set it as a content of this instance.
    drive_file.SetContentFile("export.xlsx")
    drive_file.Upload()  # Upload the file.


def export_spreadsheet():
    firebase = pyrebase.initialize_app(ast.literal_eval(os.environ['firebase_info']))
    database = firebase.database()

    teams = database.child("teams").get().each()

    team1 = teams[0].val()

    totals = ([key for key in team1['totals'].keys()], 'totals')
    l3ms = ([key for key in team1['l3ms'].keys()], 'l3ms')
    SDs = ([key for key in team1['SDs'].keys()], 'SDs')
    maxes = ([key for key in team1['maxes'].keys()], 'maxes')
    team_abilities = ([key for key in team1['team_abilities'].keys()], 'team_abilities')
    percentages = ([key for key in team1['percentages'].keys()], 'percentages')
    p75s = ([key for key in team1['p75s'].keys()], 'p75s')
    # sykes = ([key for key in team1['sykes'].keys()], 'sykes')

    timdHeader = ([key for key in team1['timds'][0]['header'].keys()], 'header')
    climb = ([key for key in team1['timds'][0]['climb'].keys()], 'climb')
    calculated = ([key for key in team1['timds'][0]['calculated'].keys()], 'calculated')

    headers = [totals, l3ms, SDs, maxes, team_abilities, percentages, p75s]  # [sykes]
    timdHeaders = [timdHeader, calculated, climb]

    wb = openpyxl.load_workbook('export.xlsx')

    for sheet in wb.worksheets:
        wb.remove(sheet)

    raw_team_sheet = wb.create_sheet('Raw Teams Export')
    raw_team_sheet.cell(row=1, column=1).value = 'Number'

    raw_timd_sheet = wb.create_sheet('Raw TIMD Export')
    raw_timd_sheet.cell(row=1, column=1).value = 'Team'
    raw_timd_sheet.cell(row=1, column=2).value = 'Match'

    current_column = 3

    for header in timdHeaders:
        for key in header[0]:
            raw_timd_sheet.cell(row=1, column=current_column).value = key
            current_column += 1

    current_column = 2

    for header in headers:
        for key in header[0]:
            raw_team_sheet.cell(row=1, column=current_column).value = key
            current_column += 1

    current_timd_row = 2
    current_team_row = 2

    for team in teams:
        team = team.val()
        current_column = 2
        raw_team_sheet.cell(row=current_team_row, column=1).value = team['teamNumber']

        for header in headers:
            for key in header[0]:
                raw_team_sheet.cell(row=current_team_row, column=current_column).value = team[header[1]][key]
                current_column += 1

        current_team_row += 1

        for timd in team['timds']:
            current_column = 3
            raw_timd_sheet.cell(row=current_timd_row, column=1).value = timd['team_number']
            raw_timd_sheet.cell(row=current_timd_row, column=2).value = timd['match_number']

            if timd['header']['noShow']:
                for key in timdHeader[0]:
                    raw_timd_sheet.cell(row=current_timd_row, column=current_column).value = timd['header'][key]
                    current_column += 1

            else:
                for header in timdHeaders:
                    for key in header[0]:
                        timd_component_header = timd[header[1]]
                        if key in timd_component_header.keys():
                            raw_timd_sheet.cell(row=current_timd_row, column=current_column).value = timd_component_header[key]
                        current_column += 1

            current_timd_row += 1

    wb.save('export.xlsx')


if __name__ == "__main__":
    export_spreadsheet()


# Download spreadsheet from excel
# Open via openpyxl
# Edit the two sheets with the updated data
# Save the spreadsheet
# Reupload the file with changes
