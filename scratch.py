from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

import openpyxl
import os

gauth = GoogleAuth()
# Try to load saved client credentials
gauth.LoadCredentialsFile("oauthcreds.txt")

if gauth.credentials is None:
    # Authenticate if they're not there
    # https://stackoverflow.com/a/55876179
    # This is what solved the issues:
    gauth.GetFlow()
    gauth.flow.params.update({'access_type': 'offline'})
    gauth.flow.params.update({'approval_prompt': 'force'})
    gauth.LocalWebserverAuth()
elif gauth.access_token_expired:
    # Refresh them if expired
    gauth.Refresh()
else:
    # Initialize the saved creds
    gauth.Authorize()
# Save the current credentials to a file
gauth.SaveCredentialsFile("oauthcreds.txt")
drive = GoogleDrive(gauth)

my_file = drive.CreateFile({'id': '1k-RZxa8-8qZXJCIhO7dtfuvYI8rqmuRR8bC4rHUm-H4'})
print('title: %s, mimeType: %s' % (my_file['title'], my_file['mimeType']))
my_file.GetContentFile('download.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

wb = openpyxl.load_workbook('download.xlsx')
raw_team_sheet = wb.create_sheet('Raw Teams Export')
raw_team_sheet.cell(row=1, column=1).value = 'Number'

my_file.SetContentFile("download.xlsx")
my_file.Upload()  # Upload the file.

os.remove("download.xlsx")